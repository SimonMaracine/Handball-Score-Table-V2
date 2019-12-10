from dataclasses import dataclass
from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THIN

from src.team import Team
import src.log

logger = src.log.get_logger(__name__)
logger.setLevel(10)

BOLD_FONT = Font(name="Calibri", size=10, bold=True)
ALIGNMENT = Alignment(horizontal="center", vertical="center")

BORDER_LEFT = Border(left=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_RIGHT = Border(right=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_TOP = Border(top=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_BOTTOM = Border(bottom=Side(border_style=BORDER_THIN, color="00000000"))

BORDER_TOP_LEFT = Border(top=Side(border_style=BORDER_THIN, color="00000000"),
                         left=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_TOP_RIGHT = Border(top=Side(border_style=BORDER_THIN, color="00000000"),
                          right=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_BOTTOM_LEFT = Border(bottom=Side(border_style=BORDER_THIN, color="00000000"),
                            left=Side(border_style=BORDER_THIN, color="00000000"))
BORDER_BOTTOM_RIGHT = Border(bottom=Side(border_style=BORDER_THIN, color="00000000"),
                             right=Side(border_style=BORDER_THIN, color="00000000"))


@dataclass
class RoundData:
    team1: Team
    team2: Team


class MatchData:

    def __init__(self):
        self.rounds: List[RoundData] = []
        self.round_time = 0
        self.timeout_time = 0
        self.suspend_time = 0

    def __str__(self):
        return f"MatchData({self.rounds})"

    def get_team_round_score(self, team: int, match_round: int) -> int:
        """Should only be called in this module"""
        assert self.rounds, "There are no rounds or the requested round doesn't exist"

        team_score = 0

        if team == 1:
            for player in self.rounds[match_round - 1].team1.players:
                team_score += player.scores
        else:
            for player in self.rounds[match_round - 1].team2.players:
                team_score += player.scores

        return team_score


def generate_report(match_data: MatchData, name: str):
    if match_data is None or len(match_data.rounds) == 0:
        logger.info("There is no data to generate")
        return

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Team 1"
    sheet2 = workbook.create_sheet("Team 2")

    _write_team_to_sheet(sheet1, match_data, team=1)
    _write_team_to_sheet(sheet2, match_data, team=2)
    _write_match_to_sheet(sheet1, match_data)

    _save_to_file(workbook, name)


def _save_to_file(workbook: Workbook, file_name: str):
    workbook.save(file_name)
    logger.info("Saved report as " + file_name)


def _write_match_to_sheet(sheet: Worksheet, match_data: MatchData):
    assert match_data.round_time != 0 and match_data.timeout_time != 0 and match_data.suspend_time != 0

    no_players = len(match_data.rounds[0].team1.players)  # no. players of first team to know the height of the table

    # Merge cells
    sheet.merge_cells(start_row=12 + no_players, start_column=1, end_row=13 + no_players, end_column=1)

    # Insert data
    sheet.cell(12 + no_players, 1, "Round")
    sheet.cell(12 + no_players, 2, "Time")
    sheet.cell(12 + no_players, 3, "Sus. time")
    sheet.cell(12 + no_players, 4, "TO time")

    sheet.cell(13 + no_players, 2, match_data.round_time)
    sheet.cell(13 + no_players, 3, match_data.suspend_time)
    sheet.cell(13 + no_players, 4, match_data.timeout_time)

    # Apply styling
    for row in range(12 + no_players, 14 + no_players):
        for col in range(1, 5):
            sheet.cell(row, col).alignment = ALIGNMENT

    sheet.cell(12 + no_players, 1).border = BORDER_TOP_LEFT
    sheet.cell(12 + no_players, 2).border = BORDER_TOP
    sheet.cell(12 + no_players, 3).border = BORDER_TOP
    sheet.cell(12 + no_players, 4).border = BORDER_TOP_RIGHT

    sheet.cell(13 + no_players, 4).border = BORDER_BOTTOM_RIGHT
    sheet.cell(13 + no_players, 3).border = BORDER_BOTTOM
    sheet.cell(13 + no_players, 2).border = BORDER_BOTTOM
    sheet.cell(13 + no_players, 1).border = BORDER_BOTTOM_LEFT


def _write_team_to_sheet(sheet: Worksheet, match_data: MatchData, team: int):
    no_rounds = len(match_data.rounds)  # no. rounds to know the width of the table
    no_players = len(match_data.rounds[0].team1.players) if team == 1 else \
            len(match_data.rounds[0].team2.players)  # no. players to know the height of the table

    # Merge top cells first table, but not rounds
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + no_rounds * 2)  # end col - 1
    sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    for col in range(1, 3 + len(match_data.rounds) * 2):
        sheet.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)

    # Merge top cells second table, but not rounds
    sheet.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2 + no_rounds * 2)  # end col - 1
    sheet.merge_cells(start_row=9, start_column=1, end_row=10, end_column=1)
    sheet.merge_cells(start_row=9, start_column=2, end_row=10, end_column=2)

    # For every match round merge the cells and enter the data
    match_round: RoundData
    for r, match_round in enumerate(match_data.rounds):  # r starts from 0
        round_team_score: int = match_data.get_team_round_score(team, r + 1)

        # Merge the corresponding cells
        sheet.merge_cells(start_row=4, start_column=3 + r * 2, end_row=4, end_column=4 + r * 2)
        sheet.merge_cells(start_row=10, start_column=3 + r * 2, end_row=10, end_column=4 + r * 2)

        # First table
        sheet.cell(4, 3 + r * 2, "Round " + str(r + 1))
        sheet.cell(5, 3 + r * 2, round_team_score)
        if team == 1:
            sheet.cell(5, 4 + r * 2, match_round.team1.time_out_requests)
        else:
            sheet.cell(5, 4 + r * 2, match_round.team2.time_out_requests)

        # Second table
        sheet.cell(10, 3 + r * 2, "Round " + str(r + 1))
        for i, player in enumerate(match_round.team1.players) if team == 1 else \
                enumerate(match_round.team2.players):
            if not player.disqualified:
                sheet.cell(11 + i, 3 + r * 2, player.scores)
                sheet.cell(11 + i, 4 + r * 2, f"{player.yellow_cards}, {player.red_cards}")
            else:
                sheet.cell(11 + i, 3 + r * 2, "n/a")
                sheet.cell(11 + i, 4 + r * 2, "n/a")

        # Insert the rest
        sheet.cell(3, 3 + r * 2, "Score")
        sheet.cell(3, 4 + r * 2, "TO reqs")
        sheet.cell(9, 3 + r * 2, "Scores")
        sheet.cell(9, 4 + r * 2, "Cards")

    # Insert the rest of the data
    sheet.cell(2, 1, "Team")
    sheet.cell(3, 1, "Name")
    sheet.cell(3, 1, "Name")
    sheet.cell(3, 2, "NO. players")
    if team == 1:
        sheet.cell(5, 1, match_data.rounds[0].team1.name)  # doesn't matter which round, because it should stay constant
        sheet.cell(5, 2, str(len(match_data.rounds[0].team1.players)))
    else:
        sheet.cell(5, 1, match_data.rounds[0].team2.name)
        sheet.cell(5, 2, str(len(match_data.rounds[0].team2.players)))

    sheet.cell(8, 1, "Players")
    sheet.cell(9, 1, "Number")
    sheet.cell(9, 2, "Name")
    if team == 1:
        for i, player in enumerate(match_data.rounds[0].team1.players):
            sheet.cell(11 + i, 1, "{:02d}".format(player.number))
            sheet.cell(11 + i, 2, player.name)
    else:
        for i, player in enumerate(match_data.rounds[0].team2.players):
            sheet.cell(11 + i, 1, "{:02d}".format(player.number))
            sheet.cell(11 + i, 2, player.name)

    # Apply styling
    ##################################################

    # Center all cells
    for row in range(2, 11 + no_players):
        for col in range(1, 3 + no_rounds * 2):
            sheet.cell(row, col).alignment = ALIGNMENT

    # Bold font to first table
    for row in range(2, 5):
        for col in range(1, 3 + no_rounds * 2):
            sheet.cell(row, col).font = BOLD_FONT

    # Bold font to second table
    for row in range(8, 11):
        for col in range(1, 3 + no_rounds * 2):
            sheet.cell(row, col).font = BOLD_FONT

    # Set a larger width for columns A and B
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 15

    # Add borders
    sheet.cell(2, 1).border = BORDER_TOP_LEFT

    sheet.cell(2, 2 + no_rounds * 2).border = BORDER_RIGHT
    sheet.cell(3, 2 + no_rounds * 2).border = BORDER_RIGHT
    sheet.cell(4, 2 + no_rounds * 2).border = BORDER_RIGHT
    sheet.cell(5, 2 + no_rounds * 2).border = BORDER_RIGHT

    for col in range(1, 3 + no_rounds * 2):
        sheet.cell(6, col).border = BORDER_BOTTOM

    sheet.cell(3, 1).border = BORDER_LEFT
    sheet.cell(5, 1).border = BORDER_LEFT

    sheet.cell(8, 1).border = BORDER_TOP_LEFT

    sheet.cell(8, 2 + no_rounds * 2).border = BORDER_RIGHT
    sheet.cell(9, 2 + no_rounds * 2).border = BORDER_RIGHT
    sheet.cell(10, 2 + no_rounds * 2).border = BORDER_RIGHT
    for row in range(11, 10 + no_players):
        sheet.cell(row, 2 + no_rounds * 2).border = BORDER_RIGHT

    sheet.cell(10 + no_players, 2 + no_rounds * 2).border = BORDER_BOTTOM_RIGHT

    for col in range(2, 2 + no_rounds * 2):
        sheet.cell(10 + no_players, col).border = BORDER_BOTTOM

    sheet.cell(10 + no_players, 1).border = BORDER_BOTTOM_LEFT

    sheet.cell(9, 1).border = BORDER_LEFT
    for row in range(11, 10 + no_players):
        sheet.cell(row, 1).border = BORDER_LEFT
    ##################################################
